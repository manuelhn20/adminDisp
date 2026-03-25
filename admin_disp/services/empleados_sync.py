"""
Sincronización de empleados desde SharePoint (Microsoft Graph) hacia la BD local.

Esta versión:
- Usa configuración cargada desde variables de entorno (no hay secretos hardcodeados).
- Logger por módulo que escribe detalles en `logs/<SYNC_LOG_FILE>`.
- Sincronización incremental: INSERT nuevos, UPDATE solo si cambian campos.
"""

from __future__ import annotations

import io
import logging
import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple, TypedDict

import msal
import openpyxl
import requests

from ..core.db import get_db_empleados

# Logger - usa configuración centralizada de app.py (escribirá a empleados_sync.log separado)
logger = logging.getLogger("admin_disp.services.empleados_sync")


class EmpleadoRow(TypedDict, total=False):
    codigo_empleado: str
    nombre_completo: str
    estado: str
    genero: str
    sucursal: str
    departamento: str
    puesto: str
    empresa: str
    usuario: str
    pasaporte: str


@dataclass
class SharePointConfig:
    tenant_id: str
    client_id: str
    client_secret: str
    site_hostname: str
    site_path: str
    drive_name: str
    filename: str
    folder_path: str = ""

    @staticmethod
    def from_env() -> "SharePointConfig":
        tenant_id = os.getenv("MS_TENANT_ID")
        client_id = os.getenv("MS_CLIENT_ID")
        client_secret = os.getenv("MS_CLIENT_SECRET")
        if not all([tenant_id, client_id, client_secret]):
            raise RuntimeError("Faltan variables de entorno requeridas: MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET")

        site_hostname = os.getenv("SP_SITE_HOSTNAME", "proimamericanos.sharepoint.com")
        site_path = os.getenv("SP_SITE_PATH", "/sites/ERPNext-Reportes")
        drive_name = os.getenv("SP_DRIVE_NAME", "Reportes Power Bi")
        filename = os.getenv("SP_FILENAME", "RPT---AT---Empleados.xlsx")
        folder_path = os.getenv("SP_FOLDER_PATH", "").strip().strip("/")

        return SharePointConfig(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
            site_hostname=site_hostname,
            site_path=site_path,
            drive_name=drive_name,
            filename=filename,
            folder_path=folder_path,
        )


class GraphClient:
    def __init__(self, cfg: SharePointConfig, session: Optional[requests.Session] = None) -> None:
        self.cfg = cfg
        self.session = session or requests.Session()

    def get_token(self) -> str:
        authority = f"https://login.microsoftonline.com/{self.cfg.tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id=self.cfg.client_id,
            authority=authority,
            client_credential=self.cfg.client_secret,
        )
        result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        token = result.get("access_token")
        if not token:
            raise RuntimeError(f"Error token: {result.get('error_description') or result}")
        return token

    def _headers(self, token: str) -> Dict[str, str]:
        return {"Authorization": f"Bearer {token}"}

    def get_site_id(self, token: str) -> str:
        url = f"https://graph.microsoft.com/v1.0/sites/{self.cfg.site_hostname}:{self.cfg.site_path}"
        r = self.session.get(url, headers=self._headers(token), timeout=60)
        r.raise_for_status()
        return r.json()["id"]

    def get_drive_id(self, token: str, site_id: str) -> str:
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
        r = self.session.get(url, headers=self._headers(token), timeout=60)
        r.raise_for_status()

        for drive in r.json().get("value", []):
            if drive.get("name") == self.cfg.drive_name:
                return drive["id"]

        available = [d.get("name", "") for d in r.json().get("value", [])]
        raise RuntimeError(f"Biblioteca/drive '{self.cfg.drive_name}' no encontrada. Disponibles: {available}")

    def get_file_item_id(self, token: str, drive_id: str) -> str:
        if self.cfg.folder_path:
            path = f"{self.cfg.folder_path}/{self.cfg.filename}"
        else:
            path = self.cfg.filename

        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{path}"
        r = self.session.get(url, headers=self._headers(token), timeout=60)
        r.raise_for_status()
        return r.json()["id"]

    def download_file_bytes(self, token: str, drive_id: str, item_id: str) -> bytes:
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
        r = self.session.get(url, headers=self._headers(token), timeout=120)
        r.raise_for_status()
        return r.content


EmpleadoRow = Dict[str, Optional[str]]


def _norm_str(value: Any, max_len: Optional[int] = None) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
    else:
        s = str(value).strip()

    if not s:
        return None
    if max_len is not None:
        s = s[:max_len]
    return s


def _detect_offset(first_data_row: Tuple[Any, ...]) -> int:
    if not first_data_row:
        return 0

    c0 = first_data_row[0] if len(first_data_row) > 0 else None
    if isinstance(c0, (int, float)):
        return 1
    return 0


def parse_empleados_excel(excel_bytes: bytes) -> List[EmpleadoRow]:
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    offset = _detect_offset(first_row or ())

    def get_cell(row: Tuple[Any, ...], idx: int) -> Any:
        real = idx + offset
        return row[real] if row and len(row) > real else None

    empleados: List[EmpleadoRow] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue

        codigo = _norm_str(get_cell(row, 0), max_len=15)
        nombre = _norm_str(get_cell(row, 1), max_len=100)

        if not codigo or not nombre:
            continue

        emp: EmpleadoRow = {
            "codigo_empleado": codigo,
            "nombre_completo": nombre,
            "estado": _norm_str(get_cell(row, 2)),
            "genero": _norm_str(get_cell(row, 3)),
            "sucursal": _norm_str(get_cell(row, 4)),
            "departamento": _norm_str(get_cell(row, 5)),
            "puesto": _norm_str(get_cell(row, 6)),
            "empresa": _norm_str(get_cell(row, 7)),
            "usuario": _norm_str(get_cell(row, 8)),
            "pasaporte": _norm_str(get_cell(row, 9), max_len=15),
        }
        empleados.append(emp)

    logger.debug("Total procesados: %s", len(empleados))
    return empleados


def obtener_empleados_sharepoint(cfg: Optional[SharePointConfig] = None) -> List[EmpleadoRow]:
    cfg = cfg or SharePointConfig.from_env()
    graph = GraphClient(cfg)
    logger.debug("Descargando empleados desde SharePoint...")
    token = graph.get_token()
    site_id = graph.get_site_id(token)
    drive_id = graph.get_drive_id(token, site_id)
    item_id = graph.get_file_item_id(token, drive_id)
    excel_bytes = graph.download_file_bytes(token, drive_id, item_id)

    empleados = parse_empleados_excel(excel_bytes)
    logger.debug("Descarga exitosa: %s empleados", len(empleados))
    return empleados


ComparableRow = Dict[str, Optional[str]]


def _row_for_compare(emp: EmpleadoRow) -> ComparableRow:
    return {
        "nombre_completo": emp.get("nombre_completo"),
        "estado": emp.get("estado"),
        "genero": emp.get("genero"),
        "sucursal": emp.get("sucursal"),
        "departamento": emp.get("departamento"),
        "puesto": emp.get("puesto"),
        "empresa": emp.get("empresa"),
        "usuario": emp.get("usuario"),
        "pasaporte": emp.get("pasaporte"),
    }


def _fetch_existing(cur) -> Dict[str, ComparableRow]:
    cur.execute(
        """
        SELECT
            codigo_empleado,
            nombre_completo,
            estado,
            genero,
            sucursal,
            departamento,
            puesto,
            empresa,
            usuario,
            pasaporte
        FROM empleados
        """
    )
    existing: Dict[str, ComparableRow] = {}
    for row in cur.fetchall() or []:
        codigo = _norm_str(row[0])
        if not codigo:
            continue
        existing[codigo] = {
            "nombre_completo": _norm_str(row[1]),
            "estado": _norm_str(row[2]),
            "genero": _norm_str(row[3]),
            "sucursal": _norm_str(row[4]),
            "departamento": _norm_str(row[5]),
            "puesto": _norm_str(row[6]),
            "empresa": _norm_str(row[7]),
            "usuario": _norm_str(row[8]),
            "pasaporte": _norm_str(row[9]),
        }
    return existing


def sincronizar_empleados(cfg: Optional[SharePointConfig] = None, only: Optional[str] = None) -> List[EmpleadoRow] | bool:
    try:
        try:
            from ..devices.service import ensure_admin_exists
        except ImportError:
            from ..cxc.service import ensure_admin_exists
        ensure_admin_exists()
        logger.info("=== INICIO SINCRONIZACIÓN ===")

        # Source control: only can be 'primary', 'alt' or None (both)
        cfg_primary = cfg or SharePointConfig.from_env()

        # Alternate file configuration
        alt_drive = os.getenv("SP_DRIVE_NAME_ALT", "Reportes Power Bi - ELMIGO")
        alt_filename = os.getenv("SP_FILENAME_ALT", "RA---EM---Empleados.xlsx")
        alt_folder = os.getenv("SP_FOLDER_PATH_ALT", "").strip().strip("/")

        cfg_alt = SharePointConfig(
            tenant_id=cfg_primary.tenant_id,
            client_id=cfg_primary.client_id,
            client_secret=cfg_primary.client_secret,
            site_hostname=cfg_primary.site_hostname,
            site_path=cfg_primary.site_path,
            drive_name=alt_drive,
            filename=alt_filename,
            folder_path=alt_folder,
        )

        empleados_primary: List[EmpleadoRow] = []
        empleados_alt: List[EmpleadoRow] = []

        # Decide which sources to fetch
        fetch_primary = only in (None, 'primary')
        fetch_alt = only in (None, 'alt')

        if fetch_primary:
            logger.info("Obteniendo empleados desde SharePoint PROIMA")
            empleados_primary = obtener_empleados_sharepoint(cfg_primary)
            logger.info("Empleados obtenidos %s de PROIMA", len(empleados_primary))

        if fetch_alt:
            try:
                logger.info("Obteniendo empleados desde SharePoint ELMIGO")
                empleados_alt = obtener_empleados_sharepoint(cfg_alt)
                logger.info("Empleados obtenidos %s de ELMIGO", len(empleados_alt))
            except Exception as e:
                logger.error("Error obteniendo empleados desde ELMIGO: %s", e)

        # Build empleados list according to 'only' and dedupe policy:
        # - If only == 'primary' -> process only primary with normal insert/update behavior
        # - If only == 'alt' -> process only alt but DO NOT update existing records (policy A): insert only
        # - If only is None -> combine primary then alt (filtering alt codes present in primary)
        if only == 'primary':
            empleados = empleados_primary
        elif only == 'alt':
            empleados = empleados_alt
        else:
            primary_codes = {e.get('codigo_empleado') for e in empleados_primary if e.get('codigo_empleado')}
            alt_filtered = [e for e in empleados_alt if e.get('codigo_empleado') and e.get('codigo_empleado') not in primary_codes]
            empleados = (empleados_primary or []) + alt_filtered

        if not empleados:
            logger.warning("Sin empleados para sincronizar")
            return []

        if only is None:
            # compute alt_filtered length used for logging
            primary_len = len(empleados_primary) if empleados_primary else 0
            alt_filtered_len = 0
            try:
                primary_codes = {e.get('codigo_empleado') for e in empleados_primary if e.get('codigo_empleado')}
                alt_filtered_len = len([e for e in empleados_alt if e.get('codigo_empleado') and e.get('codigo_empleado') not in primary_codes]) if empleados_alt else 0
            except Exception:
                alt_filtered_len = len(empleados_alt) if empleados_alt else 0
            logger.info("Total archivos - principal=%s, alterno_filtrados=%s, total_combined=%s", primary_len, alt_filtered_len, len(empleados))

        conn = get_db_empleados()
        cur = conn.get_cursor()

        sql_insert = (
            "INSERT INTO empleados "
            "(codigo_empleado, nombre_completo, estado, genero, sucursal, departamento, puesto, empresa, usuario, pasaporte) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
        )
        sql_update = (
            "UPDATE empleados SET "
            "nombre_completo = ?, estado = ?, genero = ?, sucursal = ?, "
            "departamento = ?, puesto = ?, empresa = ?, usuario = ?, pasaporte = ? "
            "WHERE codigo_empleado = ?"
        )

        inserted = updated = skipped = 0
        inserted_codes: set[str] = set()
        modified_records: List[Dict[str, Any]] = []

        try:
            existing = _fetch_existing(cur)

            for emp in empleados:
                codigo = emp.get("codigo_empleado")
                if not codigo:
                    continue

                new_row = _row_for_compare(emp)
                old_row = existing.get(codigo)

                if old_row is None:
                    if codigo in inserted_codes:
                        skipped += 1
                        continue
                    try:
                        cur.execute(
                            sql_insert,
                            (
                                codigo,
                                new_row["nombre_completo"],
                                new_row["estado"],
                                new_row["genero"],
                                new_row["sucursal"],
                                new_row["departamento"],
                                new_row["puesto"],
                                new_row["empresa"],
                                new_row["usuario"],
                                new_row["pasaporte"],
                            ),
                        )
                        inserted += 1
                        inserted_codes.add(codigo)
                    except Exception as row_e:
                        logger.error("Error insertando codigo=%s: %s", codigo, row_e)
                    continue

                # If running only the alternate source, do not overwrite existing records (policy A)
                if old_row is not None and only == 'alt':
                    skipped += 1
                    continue

                if new_row == old_row:
                    skipped += 1
                    continue

                diffs = [(k, old_row.get(k), new_row.get(k)) for k in new_row.keys() if old_row.get(k) != new_row.get(k)]
                try:
                    cur.execute(
                        sql_update,
                        (
                            new_row["nombre_completo"],
                            new_row["estado"],
                            new_row["genero"],
                            new_row["sucursal"],
                            new_row["departamento"],
                            new_row["puesto"],
                            new_row["empresa"],
                            new_row["usuario"],
                            new_row["pasaporte"],
                            codigo,
                        ),
                    )
                    updated += 1
                    if diffs:
                        modified_records.append({"codigo": codigo, "diffs": diffs})
                except Exception as row_e:
                    logger.error("Error actualizando codigo=%s: %s", codigo, row_e)

            # Sincronizar el estado en la tabla `usuarios` según el campo `usuario` del Excel
            usuarios_updated = 0
            try:
                for emp_sync in empleados:
                    usuario_raw = _norm_str(emp_sync.get("usuario"))
                    estado_emp = _norm_str(emp_sync.get("estado"))
                    if not usuario_raw:
                        continue

                    # Usar el email/usuario tal cual (incluye dominio @proimahn.com / @elmigohn.com)
                    username = usuario_raw

                    # Determinar valor BIT: 0 => inactivo, 1 => activo (default)
                    estado_bit = 1
                    if estado_emp:
                        s = estado_emp.lower()
                        if "ina" in s or "inact" in s or "inativo" in s or "inactivo" in s:
                            estado_bit = 0
                        elif "act" in s or "activo" in s:
                            estado_bit = 1

                    try:
                        cur.execute("SELECT 1 FROM usuarios WHERE username = ?", (username,))
                        if cur.fetchone():
                            cur.execute("UPDATE usuarios SET estado = ? WHERE username = ?", (estado_bit, username))
                            usuarios_updated += 1
                    except Exception as u_e:
                        logger.error("Error actualizando usuario '%s' en tabla usuarios: %s", username, u_e)
            except Exception as e_u:
                logger.error("Error sincronizando estados en usuarios: %s", e_u)

            # Confirmar cambios en ambas tablas
            conn.commit()

            if modified_records:
                logger.info("#####################################")
                logger.info("# Registros modificados")
                for m in modified_records:
                    changes = [f"{f}: '{o}' -> '{n}'" for f, o, n in m["diffs"]]
                    logger.info("%s: %s", m["codigo"], "; ".join(changes))
                logger.info("#####################################")

            if inserted_codes:
                logger.info("#####################################")
                logger.info("# Registros insertados")
                for c in sorted(inserted_codes):
                    logger.info("%s", c)
                logger.info("#####################################")

            logger.info(
                "Sincronización completada: inserted=%s, updated=%s, skipped=%s, total_input=%s",
                inserted,
                updated,
                skipped,
                len(empleados),
            )
            return empleados

        except Exception as e:
            conn.rollback()
            logger.error("Error BD: %s", e)
            raise
        finally:
            try:
                cur.close()
            except Exception as e:
                logger.debug(f"Error cerrando cursor: {e}")
            try:
                conn.close()
            except Exception as e:
                logger.debug(f"Error cerrando conexión: {e}")

    except Exception as e:
        logger.error("Error crítico: %s", e)
        return False
